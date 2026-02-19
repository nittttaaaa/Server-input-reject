from flask import Flask, request, redirect, render_template_string, send_file
from openpyxl import Workbook, load_workbook
import pandas as pd
import os
import matplotlib.pyplot as plt

app = Flask(__name__)

FILE_NAME = "reject_data.xlsx"
STATIC_FOLDER = "static"
CHART_PATH = os.path.join(STATIC_FOLDER, "chart.png")

HEADERS = [
    "Date", "Process", "Problem Type",
    "Reject Quantity", "Main Customer", "Workcenter"
]

PROCESSES = [
    "UV HOCK","VARNISH","LAMINATING","WP GERMAN","WP ESATEC",
    "WP ZHENGMAO 1 C","WP ZHENGMAO 2 C",
    "FG MATTEL","FG MASTER","FG BICHENG","FG BOBST",
    "JHOOK 5 C","JHOOK 1",
    "LINE 1","LINE 2","LINE 3","LINE 4",
    *[f"PH {i}" for i in range(1,23)],
    *[f"HS {i}" for i in range(1,7)],
    "SABLON SEMI AUTO","LINE BORONGAN",
    "FG ROLAM C",
    "VACUUM 1","VACUUM 2","VACUUM 3","VACUUM 4"
]

# ========================
# INITIAL SETUP
# ========================

if not os.path.exists(STATIC_FOLDER):
    os.makedirs(STATIC_FOLDER)

if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    wb.save(FILE_NAME)

# ========================
# HTML + BACKGROUND DESIGN
# ========================

HTML = """
<!DOCTYPE html>
<html>
<head>
<title>Reject Monitoring System</title>
<style>

body {
    margin: 0;
    font-family: Arial, sans-serif;
    background: linear-gradient(135deg, #1f4037, #99f2c8);
}

.container {
    width: 90%;
    margin: 30px auto;
    background: white;
    padding: 25px;
    border-radius: 15px;
    box-shadow: 0 10px 25px rgba(0,0,0,0.2);
}

h2 {
    text-align: center;
    color: #1f4037;
}

h3 {
    color: #2c3e50;
    margin-top: 30px;
}

input, select {
    padding: 8px;
    margin: 5px;
    border-radius: 6px;
    border: 1px solid #ccc;
}

button {
    padding: 8px 15px;
    border: none;
    border-radius: 6px;
    background-color: #1f4037;
    color: white;
    cursor: pointer;
    transition: 0.3s;
}

button:hover {
    background-color: #14532d;
}

.delete-btn {
    background-color: #e74c3c;
}

.delete-btn:hover {
    background-color: #c0392b;
}

table {
    width: 100%;
    border-collapse: collapse;
    margin-top: 15px;
}

th {
    background-color: #1f4037;
    color: white;
    padding: 10px;
}

td {
    padding: 8px;
    text-align: center;
}

tr:nth-child(even) {
    background-color: #f2f2f2;
}

.chart {
    text-align: center;
    margin-top: 20px;
}

a {
    text-decoration: none;
    font-weight: bold;
    color: #1f4037;
}

a:hover {
    text-decoration: underline;
}

hr {
    margin-top: 25px;
    margin-bottom: 25px;
}

</style>
</head>

<body>

<div class="container">

<h2>REJECT MONITORING SYSTEM</h2>

<h3>Manual Input</h3>
<form method="post">
Date <input type="date" name="date" required>
Process
<select name="process">
{% for p in processes %}
<option>{{p}}</option>
{% endfor %}
</select>
Problem <input name="problem" required>
Qty <input type="number" name="qty" required>
Customer <input name="customer" required>
Workcenter <input name="workcenter" required>
<button>Add Data</button>
</form>

<hr>

<h3>Upload Excel File</h3>
<form action="/upload" method="post" enctype="multipart/form-data">
<input type="file" name="file" accept=".xlsx" required>
<button>Upload</button>
</form>

<a href="/download">Download Excel File</a>

<form action="/delete_all" method="post" style="margin-top:10px">
<button class="delete-btn">DELETE ALL DATA</button>
</form>

<hr>

<div class="chart">
<h3>Total Reject per Process</h3>
<img src="/chart" width="80%">
</div>

<hr>

<h3>Reject Data Table</h3>
<table>
<tr>
{% for h in headers %}
<th>{{h}}</th>
{% endfor %}
<th>Action</th>
</tr>

{% for row in data %}
<tr>
{% for c in row %}
<td>{{c}}</td>
{% endfor %}
<td>
<form action="/delete/{{ loop.index0 }}" method="post">
<button class="delete-btn">Delete</button>
</form>
</td>
</tr>
{% endfor %}
</table>

</div>

</body>
</html>
"""

# ========================
# ROUTES
# ========================

@app.route("/", methods=["GET","POST"])
def index():
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    if request.method == "POST":
        try:
            qty = int(request.form["qty"])
        except:
            return "Reject Quantity must be numeric"

        ws.append([
            request.form["date"],
            request.form["process"],
            request.form["problem"],
            qty,
            request.form["customer"],
            request.form["workcenter"]
        ])
        wb.save(FILE_NAME)
        return redirect("/")

    data = list(ws.iter_rows(values_only=True))[1:]

    return render_template_string(
        HTML,
        data=data,
        headers=HEADERS,
        processes=PROCESSES
    )

@app.route("/upload", methods=["POST"])
def upload():
    file = request.files["file"]
    df = pd.read_excel(file)

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    for row in df.values.tolist():
        ws.append(row)

    wb.save(FILE_NAME)
    return redirect("/")

@app.route("/delete/<int:row_id>", methods=["POST"])
def delete_row(row_id):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.delete_rows(row_id + 2)
    wb.save(FILE_NAME)
    return redirect("/")

@app.route("/delete_all", methods=["POST"])
def delete_all():
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.delete_rows(2, ws.max_row)
    wb.save(FILE_NAME)
    return redirect("/")

@app.route("/download")
def download():
    return send_file(FILE_NAME, as_attachment=True)

@app.route("/chart")
def chart():
    df = pd.read_excel(FILE_NAME)

    if df.empty:
        plt.figure()
        plt.text(0.5, 0.5, "No Data Available", ha='center')
        plt.savefig(CHART_PATH)
        plt.close()
        return send_file(CHART_PATH, mimetype='image/png')

    df["Reject Quantity"] = pd.to_numeric(df["Reject Quantity"], errors='coerce')
    summary = df.groupby("Process")["Reject Quantity"].sum()

    plt.figure(figsize=(12,6))
    summary.plot(kind="bar")
    plt.xticks(rotation=45)
    plt.title("Total Reject per Process")
    plt.xlabel("Process")
    plt.ylabel("Total Reject")
    plt.tight_layout()
    plt.savefig(CHART_PATH)
    plt.close()

    return send_file(CHART_PATH, mimetype='image/png')

if __name__ == "__main__":
    app.run(debug=False)
